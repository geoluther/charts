#! /usr/bin/python

## use: charts_withcounts.py <infile.csv>
## writes to: "Charts_YYYY-MM-DD.xlsx"
## infile is a CSV file with column names counts of plays, structure like:
## Plays, `AlbumTitle`,`TrackTitle`, `TrackArtist`, `Artist`, `AddDate`, `AlbumId`,`Genre`

## how to make better:
## run on server, include DB call
## run on server as call that returns a download
## check for utf8 chars in CSV, convert to ascii


import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import Font
from datetime import date

# for reference
columns = ['Plays', 'AlbumTitle','TrackTitle', 'TrackArtist', 'Artist', 'AddDate', 'AlbumId','Genre']

genres = [  "Unknown", "Bluegrass", "Blues", "Cajun", "Celtic", "Classical", "Country", "Folk",
            "Gospel", "Hip Hop", "International", "Jazz", "Lounge-Schlock", "Modern",
            "R & B", "Ragtime", "Rap", "Reggae", "Rock", "Soundtrack", "Space", "Spoken Word",
			"Techno", "Zydeco" ]

# infile = "SoundExchangePlaylist_033117_names.csv"
infile = sys.argv[1]
print "Reading: ", infile


wb = Workbook()

d = date.today()
outfile = 'Charts_{}.xlsx'.format(d)

sheets = [ wb.create_sheet(title=g) for g in genres ]

## remove the default sheet
wb.remove_sheet( wb['Sheet'] )

## write headers to sheets
header = ['Plays', 'Artist', 'Track','Album', 'Add Date', 'Album Id','Genre']

ft = Font(bold=True)

# create header row on sheets
for g in genres:
	ws = wb[g]
	ws.append(header)

# process csv file
with open(infile, 'rb') as f:
	x = 0
	reader = csv.DictReader(f)

	for row in reader:
		x += 1

		# merge & clean Artist, TrackArtist cells
		artist = row['TrackArtist'] + row['Artist']
		artist = artist.replace("NULL","")

		# build data row, convert Plays and AlbumId string fields to int.
		# should be in same order as header
		data = [ int(row['Plays']), artist, row['TrackTitle'], row['AlbumTitle'],
		row['AddDate'], int(row['AlbumId']), row['Genre'] ]

		# cleanup genre fields
		genre = row["Genre"]
		if genre not in genres:
			genre = "Unknown"

		wb[genre].append(data)


print "rows read >> ", x
print "writing to >> ", outfile

# freeze header row, try to add bold style
for g in genres:
	ws = wb[g]
	c = ws['A2']
	ws.freeze_panes = c
	row = ws.row_dimensions[1]
	row.font = Font(bold=True) # not working


wb.save(filename = outfile)

## sql query to generate csv

# query = """SELECT Count(*) as Plays, AlbumTitle, TrackTitle, TrackArtist, Artist, AddDate, AlbumId, Genre
# FROM SoundExchangePlaylist
# WHERE StartDateTime between '2017-02-28 00:00:01' and NOW()
# GROUP BY TrackTitle, AlbumTitle
# ORDER BY Plays DESC;"""

